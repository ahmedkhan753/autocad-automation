"""
Export Module
Generates Excel (.xlsx) and CSV output files with room schedule data.
"""

import csv
import logging
import os
from datetime import datetime
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger("autocad_extractor.exporter")

BASE_DIR = Path(__file__).resolve().parent.parent

# Styling constants
HEADER_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, size=11)
DATA_FONT = Font(name="Calibri", size=11)
SUM_FONT = Font(name="Calibri", bold=True, size=11, color="1F4E79")
SUM_FILL = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Column headers for the Room Schedule sheet
HEADERS = ["#", "Room Name", "Area (sqm)", "Perimeter (m)", "Layer", "Notes"]


def export_results(
    rooms: List,
    output_dir: str = None,
) -> dict:
    """
    Export room data to both Excel and CSV files.

    Args:
        rooms: List of RoomData objects from the calculator
        output_dir: Directory to save output files (default: project output/)

    Returns:
        Dict with paths: {"excel": "path/to/xlsx", "csv": "path/to/csv"}

    Raises:
        OSError: If output directory cannot be created
        RuntimeError: If export fails
    """
    if output_dir is None:
        output_dir = str(BASE_DIR / "output")

    # Ensure output directory exists
    os.makedirs(output_dir, exist_ok=True)

    # Generate timestamped filenames
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = f"room_schedule_{timestamp}.xlsx"
    csv_filename = f"room_schedule_{timestamp}.csv"
    excel_path = os.path.join(output_dir, excel_filename)
    csv_path = os.path.join(output_dir, csv_filename)

    try:
        _export_excel(rooms, excel_path)
        logger.info(f"Excel file saved: {excel_path}")
    except Exception as e:
        logger.error(f"Failed to export Excel: {e}")
        raise RuntimeError(f"Excel export failed: {e}")

    try:
        _export_csv(rooms, csv_path)
        logger.info(f"CSV file saved: {csv_path}")
    except Exception as e:
        logger.error(f"Failed to export CSV: {e}")
        raise RuntimeError(f"CSV export failed: {e}")

    return {
        "excel": excel_path,
        "csv": csv_path,
        "excel_filename": excel_filename,
        "csv_filename": csv_filename,
    }


def _export_excel(rooms: List, filepath: str):
    """
    Create a formatted Excel workbook with Room Schedule and Raw Data sheets.
    """
    wb = Workbook()

    # ── Sheet 1: Room Schedule ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Room Schedule"

    # Write headers
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # Write data rows
    for row_idx, room in enumerate(rooms, 2):
        row_data = [
            row_idx - 1,           # # (sequential number)
            room.room_name,        # Room Name
            room.area_sqm,         # Area (sqm)
            room.perimeter_m,      # Perimeter (m)
            room.layer,            # Layer
            room.notes,            # Notes
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if col_idx in (3, 4):  # Numeric columns
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")

    # SUM row
    sum_row = len(rooms) + 2
    ws.cell(row=sum_row, column=1, value="").border = THIN_BORDER
    sum_label = ws.cell(row=sum_row, column=2, value="TOTAL")
    sum_label.font = SUM_FONT
    sum_label.fill = SUM_FILL
    sum_label.border = THIN_BORDER

    # Area SUM formula
    area_sum = ws.cell(
        row=sum_row, column=3,
        value=f"=SUM(C2:C{sum_row - 1})"
    )
    area_sum.font = SUM_FONT
    area_sum.fill = SUM_FILL
    area_sum.number_format = "#,##0.00"
    area_sum.border = THIN_BORDER
    area_sum.alignment = Alignment(horizontal="right")

    # Perimeter SUM formula
    perim_sum = ws.cell(
        row=sum_row, column=4,
        value=f"=SUM(D2:D{sum_row - 1})"
    )
    perim_sum.font = SUM_FONT
    perim_sum.fill = SUM_FILL
    perim_sum.number_format = "#,##0.00"
    perim_sum.border = THIN_BORDER
    perim_sum.alignment = Alignment(horizontal="right")

    # Fill remaining SUM row cells
    for col_idx in (5, 6):
        cell = ws.cell(row=sum_row, column=col_idx, value="")
        cell.fill = SUM_FILL
        cell.border = THIN_BORDER

    # Auto-fit column widths
    _autofit_columns(ws)

    # ── Sheet 2: Raw Data ───────────────────────────────────────────────
    ws_raw = wb.create_sheet("Raw Data")
    raw_headers = ["Room Name", "Layer", "Vertex Index", "X", "Y"]
    for col_idx, header in enumerate(raw_headers, 1):
        cell = ws_raw.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    raw_row = 2
    for room in rooms:
        for v_idx, (x, y) in enumerate(room.polygon_vertices):
            ws_raw.cell(row=raw_row, column=1, value=room.room_name).border = THIN_BORDER
            ws_raw.cell(row=raw_row, column=2, value=room.layer).border = THIN_BORDER
            ws_raw.cell(row=raw_row, column=3, value=v_idx + 1).border = THIN_BORDER
            x_cell = ws_raw.cell(row=raw_row, column=4, value=round(x, 4))
            x_cell.border = THIN_BORDER
            x_cell.number_format = "#,##0.0000"
            y_cell = ws_raw.cell(row=raw_row, column=5, value=round(y, 4))
            y_cell.border = THIN_BORDER
            y_cell.number_format = "#,##0.0000"
            raw_row += 1

    _autofit_columns(ws_raw)

    wb.save(filepath)


def _export_csv(rooms: List, filepath: str):
    """Create a CSV file with room schedule data."""
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(HEADERS)

        total_area = 0.0
        total_perimeter = 0.0

        for idx, room in enumerate(rooms, 1):
            writer.writerow([
                idx,
                room.room_name,
                room.area_sqm,
                room.perimeter_m,
                room.layer,
                room.notes,
            ])
            total_area += room.area_sqm
            total_perimeter += room.perimeter_m

        # SUM row
        writer.writerow([
            "",
            "TOTAL",
            round(total_area, 2),
            round(total_perimeter, 2),
            "",
            "",
        ])


def _autofit_columns(ws):
    """Auto-fit column widths based on content length."""
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)

        for row in ws.iter_rows(
            min_col=col_idx, max_col=col_idx,
            min_row=1, max_row=ws.max_row,
        ):
            for cell in row:
                if cell.value:
                    cell_length = len(str(cell.value))
                    max_length = max(max_length, cell_length)

        # Add padding
        adjusted_width = min(max_length + 4, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
